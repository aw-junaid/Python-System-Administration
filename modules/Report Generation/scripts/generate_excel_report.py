#!/usr/bin/env python3
"""
generate_excel_report.py

Generates a formatted Excel workbook (.xlsx) with styled headers,
a formula column (computed with real Excel formulas, not just
static values), and an embedded bar chart. Requires 'openpyxl'.

Usage:
    python generate_excel_report.py
    python generate_excel_report.py --input data.csv --output report.xlsx --title "Sales Report"

Expected output:
    An .xlsx file with one sheet named "Report" containing a styled
    table, a "Total" formula row using SUM(), and a bar chart plotted
    from the numeric column. Opens in Excel, LibreOffice Calc, or
    Google Sheets (upload/import).
"""

import argparse
import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency 'openpyxl'. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def load_sample_data():
    headers = ["Product", "Region", "Units Sold", "Revenue"]
    rows = [
        ["Widget A", "North", 120, 3600],
        ["Widget B", "South", 95, 2850],
        ["Widget C", "East", 150, 6000],
        ["Widget D", "West", 80, 2400],
        ["Widget E", "North", 60, 1800],
    ]
    return headers, rows


def load_csv_data(path):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        raise ValueError("CSV file is empty.")
    headers = rows[0]
    data_rows = []
    for r in rows[1:]:
        converted = []
        for cell in r:
            try:
                converted.append(float(cell) if "." in cell else int(cell))
            except ValueError:
                converted.append(cell)
        data_rows.append(converted)
    return headers, data_rows


def find_numeric_column(headers, rows):
    """Return index of the last numeric column, used for the chart + SUM formula."""
    for idx in range(len(headers) - 1, -1, -1):
        if rows and isinstance(rows[0][idx], (int, float)):
            return idx
    return None


def build_workbook(headers, rows, title, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(*(Side(style="thin", color="CBD5E1"),) * 4)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True, color="1D4ED8")
    ws.row_dimensions[1].height = 24

    # Header row
    header_row_idx = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for r_offset, row in enumerate(rows, start=1):
        for c_offset, value in enumerate(row, start=1):
            cell = ws.cell(row=header_row_idx + r_offset, column=c_offset, value=value)
            cell.border = thin_border

    numeric_col = find_numeric_column(headers, rows)

    # Totals row using a real Excel formula
    total_row_idx = header_row_idx + len(rows) + 1
    ws.cell(row=total_row_idx, column=1, value="TOTAL").font = Font(bold=True)
    if numeric_col is not None:
        col_letter = get_column_letter(numeric_col + 1)
        first_data_row = header_row_idx + 1
        last_data_row = header_row_idx + len(rows)
        formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        total_cell = ws.cell(row=total_row_idx, column=numeric_col + 1, value=formula)
        total_cell.font = Font(bold=True)

    # Column widths
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(header)) + 4)

    # Chart
    if numeric_col is not None and len(rows) > 1:
        chart = BarChart()
        chart.title = f"{headers[numeric_col]} by {headers[0]}"
        chart.y_axis.title = headers[numeric_col]
        chart.x_axis.title = headers[0]

        data_ref = Reference(
            ws, min_col=numeric_col + 1, max_col=numeric_col + 1,
            min_row=header_row_idx, max_row=header_row_idx + len(rows)
        )
        cats_ref = Reference(
            ws, min_col=1, max_col=1,
            min_row=header_row_idx + 1, max_row=header_row_idx + len(rows)
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 16
        chart.height = 9
        ws.add_chart(chart, get_column_letter(len(headers) + 2) + "3")

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate a formatted Excel report with a chart from CSV data.")
    parser.add_argument("--input", "-i", help="Path to input CSV file. If omitted, sample data is used.")
    parser.add_argument("--output", "-o", default="report.xlsx", help="Path to output XLSX file (default: report.xlsx)")
    parser.add_argument("--title", "-t", default="Sales Performance Report", help="Report title")
    args = parser.parse_args()

    if args.input:
        if not os.path.isfile(args.input):
            print(f"Error: input file not found: {args.input}", file=sys.stderr)
            sys.exit(1)
        headers, rows = load_csv_data(args.input)
    else:
        headers, rows = load_sample_data()
        print("No --input provided, using built-in sample data.")

    build_workbook(headers, rows, args.title, args.output)

    print(f"Excel report written to: {os.path.abspath(args.output)}")
    print("Open it in Excel or LibreOffice Calc. Contains a live SUM() formula and a bar chart.")


if __name__ == "__main__":
    main()
