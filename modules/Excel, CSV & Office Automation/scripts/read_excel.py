#!/usr/bin/env python3
"""
read_excel.py
=============
Open an .xlsx file and iterate over every sheet, printing a summary of
each sheet's dimensions, header row, and a preview of the data.

Usage
-----
    python read_excel.py --file sample.xlsx
    python read_excel.py --file sample.xlsx --max-rows 5

If --file is omitted, the script generates a small sample workbook
(sample.xlsx) in the current directory so you can see the script work
end-to-end without needing your own file.

Expected output
----------------
For each sheet in the workbook you will see:
    - Sheet name
    - Number of rows / columns
    - The header row (assumed to be row 1)
    - Up to --max-rows data rows printed as tuples

Requirements
------------
    pip install openpyxl
"""

import argparse
import os
import sys

from openpyxl import Workbook, load_workbook


def create_sample_workbook(path: str) -> None:
    """Create a small multi-sheet workbook so the script is runnable out of the box."""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["ID", "Name", "Department", "Salary"])
    ws1.append([1, "Alice Khan", "Engineering", 95000])
    ws1.append([2, "Bilal Ahmed", "Sales", 62000])
    ws1.append([3, "Sara Malik", "Marketing", 71000])

    ws2 = wb.create_sheet("Inventory")
    ws2.append(["SKU", "Item", "Quantity", "Unit Price"])
    ws2.append(["SKU-001", "Laptop", 12, 850.00])
    ws2.append(["SKU-002", "Monitor", 30, 150.00])

    wb.save(path)
    print(f"[info] No --file given, created a sample workbook at: {path}")


def read_excel(path: str, max_rows: int) -> None:
    if not os.path.exists(path):
        print(f"[error] File not found: {path}", file=sys.stderr)
        sys.exit(1)

    # data_only=True returns cached formula results instead of formula strings
    wb = load_workbook(path, data_only=True)

    print(f"Workbook: {path}")
    print(f"Sheets found: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {ws.dimensions}  (rows={ws.max_row}, cols={ws.max_column})")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            print("  (sheet is empty)\n")
            continue

        print(f"Header: {header}")
        print(f"Preview (up to {max_rows} rows):")
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                print(f"  ... ({ws.max_row - 1 - max_rows} more rows not shown)")
                break
            print(f"  {row}")
        print()


def main():
    parser = argparse.ArgumentParser(description="Read and preview an Excel file sheet by sheet.")
    parser.add_argument("--file", default=None, help="Path to the .xlsx file to read")
    parser.add_argument("--max-rows", type=int, default=10, help="Max data rows to preview per sheet")
    args = parser.parse_args()

    path = args.file
    if path is None:
        path = "sample.xlsx"
        create_sample_workbook(path)

    read_excel(path, args.max_rows)


if __name__ == "__main__":
    main()
