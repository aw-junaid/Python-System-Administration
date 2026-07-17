#!/usr/bin/env python3
"""
csv_to_excel.py
================
Wrap a raw CSV file into a formatted .xlsx workbook for stakeholders
(bold header, auto column widths, frozen header row).

Usage
-----
    python csv_to_excel.py --file data.csv --output data.xlsx

If --file is omitted, a small sample CSV is generated first.

Expected output
----------------
A .xlsx file containing the CSV data, with row 1 bolded, columns
auto-sized, and the header row frozen so it stays visible while
scrolling.

Requirements
------------
    pip install openpyxl
"""

import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def create_sample_csv(path: str) -> None:
    content = (
        "Invoice,Client,Amount,Status\n"
        "INV-1001,Falcon Traders,1250.00,Paid\n"
        "INV-1002,Orion Textiles,890.50,Pending\n"
        "INV-1003,Nimbus Foods,430.00,Overdue\n"
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"[info] No --file given, created a sample CSV at: {path}")


def csv_to_excel(csv_path: str, output: str, sheet_name: str) -> None:
    if not os.path.exists(csv_path):
        print(f"[error] File not found: {csv_path}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        row_count = 0
        for row in reader:
            ws.append(row)
            row_count += 1

    if row_count > 0:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    wb.save(output)
    print(f"[success] Converted '{csv_path}' -> '{output}'")
    print(f"          Rows written (incl. header): {row_count}")


def main():
    parser = argparse.ArgumentParser(description="Convert a CSV file into a formatted Excel workbook.")
    parser.add_argument("--file", default=None, help="Path to the input CSV file")
    parser.add_argument("--output", default="converted.xlsx", help="Path for the output .xlsx file")
    parser.add_argument("--sheet", default="Sheet1", help="Name for the resulting sheet")
    args = parser.parse_args()

    path = args.file
    if path is None:
        path = "sample.csv"
        create_sample_csv(path)

    csv_to_excel(path, args.output, args.sheet)


if __name__ == "__main__":
    main()
