#!/usr/bin/env python3
"""
format_excel.py
================
Apply visual formatting to an Excel worksheet: bold/colored header font,
auto-sized column widths, and conditional formatting that highlights
cells in a numeric column based on their value (e.g. highlight sales
figures below a threshold in red, above in green).

Usage
-----
    python format_excel.py --file sample.xlsx --sheet Sheet1 --value-column D --threshold 50

If --file is omitted, a sample workbook (sample.xlsx) is generated so
the script can be run immediately.

Expected output
----------------
The target .xlsx file is modified in place (or written to --output if
given): header row becomes bold with a colored fill, columns are
auto-widened, and cells in --value-column get a red/green fill
depending on whether they are below/above --threshold.

Requirements
------------
    pip install openpyxl
"""

import argparse
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def create_sample_workbook(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Region", "Rep", "Month", "Sales"])
    ws.append(["North", "Ali Raza", "June", 42])
    ws.append(["South", "Mehak Sohail", "June", 78])
    ws.append(["East", "Usman Tariq", "June", 25])
    ws.append(["West", "Hina Baig", "June", 91])
    wb.save(path)
    print(f"[info] No --file given, created a sample workbook at: {path}")


def format_excel(path: str, output: str, sheet_name: str, value_column: str, threshold: float) -> None:
    if not os.path.exists(path):
        print(f"[error] File not found: {path}")
        return

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # --- Header formatting: bold white text on a blue fill ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- Auto-size columns ---
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # --- Conditional formatting on the chosen column (excluding header) ---
    last_row = ws.max_row
    data_range = f"{value_column}2:{value_column}{last_row}"

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="lessThan", formula=[str(threshold)], fill=red_fill),
    )
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="greaterThanOrEqual", formula=[str(threshold)], fill=green_fill),
    )

    wb.save(output)
    print(f"[success] Formatted workbook saved to: {output}")
    print(f"          Header styled, columns auto-sized, conditional formatting applied to {data_range}")
    print(f"          Rule: < {threshold} -> red, >= {threshold} -> green")


def main():
    parser = argparse.ArgumentParser(description="Apply formatting and conditional formatting to an Excel sheet.")
    parser.add_argument("--file", default=None, help="Path to the .xlsx file to format")
    parser.add_argument("--output", default=None, help="Path to save the formatted file (defaults to overwriting --file)")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name to format")
    parser.add_argument("--value-column", default="D", help="Column letter containing numeric values, e.g. D")
    parser.add_argument("--threshold", type=float, default=50, help="Threshold used for conditional formatting")
    args = parser.parse_args()

    path = args.file
    if path is None:
        path = "sample.xlsx"
        create_sample_workbook(path)

    output = args.output or path
    format_excel(path, output, args.sheet, args.value_column, args.threshold)


if __name__ == "__main__":
    main()
