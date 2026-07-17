#!/usr/bin/env python3
"""
excel_to_csv.py
================
Extract tabular data from an Excel workbook into plain CSV files, one
CSV per sheet.

Usage
-----
    python excel_to_csv.py --file report.xlsx --output-dir csv_output
    python excel_to_csv.py --file report.xlsx --sheet Summary --output summary.csv

If --file is omitted, a small sample workbook is generated first.

Expected output
----------------
- Without --sheet: one CSV file per worksheet, written to --output-dir
  (default "csv_output"), named after each sheet.
- With --sheet: a single CSV file for that sheet, written to --output
  (default "<sheet>.csv").

Requirements
------------
    pip install openpyxl
"""

import argparse
import csv
import os

from openpyxl import Workbook, load_workbook


def create_sample_workbook(path: str) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["ID", "Name", "Department"])
    ws1.append([1, "Ayesha Siddiqui", "Finance"])
    ws1.append([2, "Kamran Sheikh", "Operations"])

    ws2 = wb.create_sheet("Budget")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", 2000])
    ws2.append(["Utilities", 350])

    wb.save(path)
    print(f"[info] No --file given, created a sample workbook at: {path}")


def sheet_to_csv(ws, out_path: str) -> int:
    row_count = 0
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
            row_count += 1
    return row_count


def excel_to_csv(path: str, output_dir: str, single_sheet: str, single_output: str) -> None:
    if not os.path.exists(path):
        print(f"[error] File not found: {path}")
        return

    wb = load_workbook(path, data_only=True)

    if single_sheet:
        if single_sheet not in wb.sheetnames:
            print(f"[error] Sheet '{single_sheet}' not found. Available: {wb.sheetnames}")
            return
        out_path = single_output or f"{single_sheet}.csv"
        rows = sheet_to_csv(wb[single_sheet], out_path)
        print(f"[success] Sheet '{single_sheet}' -> '{out_path}' ({rows} rows)")
        return

    os.makedirs(output_dir, exist_ok=True)
    for sheet_name in wb.sheetnames:
        safe_name = "".join(c if c.isalnum() else "_" for c in sheet_name)
        out_path = os.path.join(output_dir, f"{safe_name}.csv")
        rows = sheet_to_csv(wb[sheet_name], out_path)
        print(f"[success] Sheet '{sheet_name}' -> '{out_path}' ({rows} rows)")


def main():
    parser = argparse.ArgumentParser(description="Extract Excel sheets into CSV files.")
    parser.add_argument("--file", default=None, help="Path to the .xlsx file to read")
    parser.add_argument("--output-dir", default="csv_output", help="Directory for per-sheet CSVs")
    parser.add_argument("--sheet", default=None, help="Extract only this sheet instead of all sheets")
    parser.add_argument("--output", default=None, help="Output CSV path when --sheet is used")
    args = parser.parse_args()

    path = args.file
    if path is None:
        path = "sample.xlsx"
        create_sample_workbook(path)

    excel_to_csv(path, args.output_dir, args.sheet, args.output)


if __name__ == "__main__":
    main()
