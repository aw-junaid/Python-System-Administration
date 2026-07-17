#!/usr/bin/env python3
"""
write_excel.py
==============
Create a new .xlsx spreadsheet from scratch (or by appending to an
existing workbook used as a template).

Usage
-----
    python write_excel.py --output report.xlsx
    python write_excel.py --output report.xlsx --template existing.xlsx --sheet "Data"

Expected output
----------------
A new (or updated) .xlsx file on disk containing a header row and the
sample rows written by this script, with basic formatting applied
(bold header, auto column width). Prints the path of the file written.

Requirements
------------
    pip install openpyxl
"""

import argparse
import os
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SAMPLE_DATA = [
    ["Order ID", "Customer", "Product", "Quantity", "Total"],
    [1001, "Zara Ahmed", "Wireless Mouse", 3, 45.00],
    [1002, "Hamza Iqbal", "Keyboard", 1, 30.00],
    [1003, "Noor Fatima", "USB-C Hub", 2, 58.00],
]


def write_excel(output: str, template: Optional[str], sheet: str) -> None:
    if template and os.path.exists(template):
        wb = load_workbook(template)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        print(f"[info] Using template '{template}', writing to sheet '{sheet}'")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet

    for row in SAMPLE_DATA:
        ws.append(row)

    # Bold the header row (row 1)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto-fit column widths based on max content length
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    wb.save(output)
    print(f"[success] Workbook written to: {output}")
    print(f"          Sheet: '{ws.title}', rows written: {len(SAMPLE_DATA)}")


def main():
    parser = argparse.ArgumentParser(description="Write a formatted Excel spreadsheet.")
    parser.add_argument("--output", default="output.xlsx", help="Path for the resulting .xlsx file")
    parser.add_argument("--template", default=None, help="Optional existing .xlsx file to write into")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name to write to")
    args = parser.parse_args()

    write_excel(args.output, args.template, args.sheet)


if __name__ == "__main__":
    main()
