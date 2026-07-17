#!/usr/bin/env python3
"""
generate_reports.py
====================
Assemble a multi-sheet analytical Excel report from one or more CSV
files. Each CSV becomes its own formatted sheet, plus a "Summary"
sheet is auto-generated with row counts and basic numeric stats
(sum/average) for each source file.

Usage
-----
    python generate_reports.py --inputs sales.csv inventory.csv --output report.xlsx

If --inputs is omitted, sample CSV data is generated in-memory so the
script produces a working multi-sheet report immediately.

Expected output
----------------
A single .xlsx file with:
    - One sheet per input CSV (named after the file), header bolded
    - A "Summary" sheet listing each source sheet's row count and,
      for the first numeric column found, its sum and average

Requirements
------------
    pip install openpyxl
"""

import argparse
import csv
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SAMPLE_CSVS = {
    "sales": "Region,Rep,Sales\nNorth,Ali Raza,4200\nSouth,Mehak Sohail,7800\nEast,Usman Tariq,2500\n",
    "inventory": "SKU,Item,Quantity\nSKU-001,Laptop,12\nSKU-002,Monitor,30\nSKU-003,Keyboard,50\n",
}


def read_csv_rows(path_or_text: str, is_path: bool):
    if is_path:
        with open(path_or_text, newline="", encoding="utf-8") as f:
            return list(csv.reader(f))
    else:
        return list(csv.reader(io.StringIO(path_or_text)))


def sheet_name_from_path(path: str) -> str:
    name = os.path.splitext(os.path.basename(path))[0]
    return name[:31]  # Excel sheet name limit


def try_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def generate_report(inputs, output: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # start with no default sheet

    summary_rows = []  # (sheet_name, row_count, numeric_col, total, average)

    for label, rows in inputs:
        if not rows:
            continue
        header, data_rows = rows[0], rows[1:]

        ws = wb.create_sheet(title=label)
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        numeric_col_idx = None
        numeric_values = []
        for row in data_rows:
            ws.append(row)
            if numeric_col_idx is None:
                for idx, val in enumerate(row):
                    if try_float(val) is not None:
                        numeric_col_idx = idx
                        break
            if numeric_col_idx is not None and numeric_col_idx < len(row):
                num = try_float(row[numeric_col_idx])
                if num is not None:
                    numeric_values.append(num)

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

        numeric_col_name = header[numeric_col_idx] if numeric_col_idx is not None else "N/A"
        total = sum(numeric_values) if numeric_values else 0
        avg = (total / len(numeric_values)) if numeric_values else 0
        summary_rows.append((label, len(data_rows), numeric_col_name, total, avg))

    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws.append(["Sheet", "Row Count", "Numeric Column", "Total", "Average"])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    for row in summary_rows:
        summary_ws.append([row[0], row[1], row[2], round(row[3], 2), round(row[4], 2)])
    for col_idx, col_cells in enumerate(summary_ws.columns, start=1):
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    wb.save(output)
    print(f"[success] Report written to: {output}")
    print(f"          Sheets created: {wb.sheetnames}")


def main():
    parser = argparse.ArgumentParser(description="Assemble a multi-sheet analytical report from CSV files.")
    parser.add_argument("--inputs", nargs="*", default=None, help="Paths to input CSV files")
    parser.add_argument("--output", default="analytical_report.xlsx", help="Path for the output .xlsx report")
    args = parser.parse_args()

    if args.inputs:
        prepared = [(sheet_name_from_path(p), read_csv_rows(p, is_path=True)) for p in args.inputs]
    else:
        print("[info] No --inputs given, using built-in sample sales/inventory data")
        prepared = [(label, read_csv_rows(text, is_path=False)) for label, text in SAMPLE_CSVS.items()]

    generate_report(prepared, args.output)


if __name__ == "__main__":
    main()
